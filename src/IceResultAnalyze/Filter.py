#!/usr/bin/python
from openpyxl import load_workbook
import sys
import os
import re
import json
import pprint

def open_data(name):
    data = []
    try:
      wb = load_workbook(filename=name)
    except (IOError, ValueError) as e:
      print e
    ws = wb['rawdata_exec']
    #find target col
    iresult = ws.min_column
    isha1 = iresult
    isha256 = iresult
    igroup = iresult
    itx = iresult
    ifiletype = iresult

    for c in range(ws.min_column,ws.max_column):
        val = ws.cell(row = ws.min_row,column = c).value
        if val == u'dt_ddan_win7x64_atseoff':
            iresult = c
        if val == u'sha1':
            isha1 = c
        if val == u'sha256':
            isha256 = c
        if val == u'group':
            igroup = c
        if val == u'dt_trendx':
            itx = c
        if val == u'ddan_win7x64_truefiletype':
            ifiletype = c

    for r in range(2,ws.max_row):
        record={}
        record['sha1'] = ws.cell(row = r, column = isha1).value
        record['sha256'] = ws.cell(row = r, column = isha256).value
        record['result'] = ws.cell(row = r, column = iresult).value
        record['group'] = ws.cell(row = r, column = igroup).value
        record['trendx'] = ws.cell(row = r, column = itx).value
        record['filetype'] = ws.cell(row=r, column= ifiletype).value

        if record['filetype'] is not None and 'DLL' not in record['filetype']:
            print record['filetype']
            data.append(record['sha1'])

    return data

def open_all_data(path):
    alldata = []
    for root,dirs,files in os.walk(path):
        for f in files:
            filepath = os.path.join(root,f)
            if re.match(r'.+ddan-[0-9]{4}-[0-9]{2}-[0-9]{2}.xlsx$',filepath):
                alldata.append(open_data(filepath))
    return alldata

def get_tx_undetect_list(data):
    return filter(lambda record:record['trendx']==0,data)

def get_fn_list(data):
    return filter(lambda record:record['result']==0,data)
def get_all_tx_FN_list(data):
    allfnlist = {}
    for k in data:
        allfnlist[k] = get_tx_undetect_list(data[k])
    return allfnlist

def get_all_fn_list(data):
    allfnlist = {}
    for k in data:
        allfnlist[k] = get_fn_list(data[k])
    return allfnlist

def count_fn(data):
    c = 0
    for i in data:
        pprint.pprint(data[i])
        c = c + len(data[i])
    return c

if __name__ == '__main__':
    targetdir = sys.argv[1]
    outfile = sys.argv[2]
    d = open_all_data(targetdir)
    #l = get_all_fn_list(d)
    #l = get_all_tx_FN_list(l)
    resfile = os.path.join(targetdir,outfile)
    json.dump(d,open(resfile,'w'))
    #print "summury: fn count: %d\n" % count_fn(l)
